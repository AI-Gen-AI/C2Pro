
import openpyxl
import sys

file_path = "docs/assets/samples/España/P-009245  Planta Bionergía de Campillos Málaga/Documentación proyecto/Presupuesto cero/20150128 RD/Presupuesto cero.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # Let's try to find a sheet that looks like a summary
    # Common names: 'Resumen', 'Presupuesto', 'TOTAL', 'Cero'
    target_sheet = None
    for name in wb.sheetnames:
        if 'Resumen' in name or 'Cero' in name or 'Presupuesto' in name:
            target_sheet = name
            break
    
    if not target_sheet:
        target_sheet = wb.sheetnames[0]
        
    print(f"Reading sheet: {target_sheet}")
    ws = wb[target_sheet]
    
    for row in ws.iter_rows(max_row=50):
        row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
        if any(row_data):
            print("\t".join(row_data))

except Exception as e:
    print(f"Error: {e}")
