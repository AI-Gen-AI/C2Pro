"""
C2Pro - Excel Parser Module

This module provides functionality to parse schedule and budget data from Excel files
using the openpyxl library. It assumes standard formats for these documents.
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


class ExcelParsingError(Exception):
    """Custom exception for Excel parsing errors."""

    pass


def parse_schedule_from_excel(file_path: Path) -> list[dict[str, Any]]:
    """
    Parses schedule data from a standard Excel file format.

    Assumes the first sheet contains the schedule with a header row.
    Expected columns (case-insensitive): 'Task', 'Start Date', 'End Date', 'Duration'.

    Args:
        file_path: The path to the Excel file (.xlsx).

    Returns:
        A list of dictionaries, where each dictionary represents a schedule task.

    Raises:
        ExcelParsingError: If the file cannot be opened or does not match the expected format.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        headers = [cell.value.lower() if cell.value else "" for cell in sheet[1]]

        # Define expected headers
        expected_headers = ["task", "start date", "end date"]
        if not all(h in headers for h in expected_headers):
            raise ExcelParsingError(
                f"Missing one or more required headers: {', '.join(expected_headers)}"
            )

        schedule_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            # Filter out empty rows
            if any(row_data.values()):
                schedule_data.append(
                    {
                        "task": row_data.get("task"),
                        "start_date": row_data.get("start date"),
                        "end_date": row_data.get("end date"),
                        "duration": row_data.get("duration"),
                    }
                )
        return schedule_data

    except (InvalidFileException, FileNotFoundError) as e:
        raise ExcelParsingError(f"Failed to open or read Excel file: {e}")
    except Exception as e:
        raise ExcelParsingError(f"An unexpected error occurred during Excel schedule parsing: {e}")


def parse_budget_from_excel(file_path: Path) -> list[dict[str, Any]]:
    """
    Parses budget data from a standard Excel file format.

    Assumes the first sheet contains the budget with a header row.
    Expected columns (case-insensitive): 'Item', 'Quantity', 'Unit Price', 'Total'.

    Args:
        file_path: The path to the Excel file (.xlsx).

    Returns:
        A list of dictionaries, where each dictionary represents a budget line item.

    Raises:
        ExcelParsingError: If the file cannot be opened or does not match the expected format.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        headers = [cell.value.lower() if cell.value else "" for cell in sheet[1]]

        # Define expected headers
        expected_headers = ["item", "quantity", "unit price", "total"]
        if not all(h in headers for h in expected_headers):
            raise ExcelParsingError(
                f"Missing one or more required headers: {', '.join(expected_headers)}"
            )

        budget_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            if any(row_data.values()):
                budget_data.append(
                    {
                        "item": row_data.get("item"),
                        "quantity": row_data.get("quantity"),
                        "unit_price": row_data.get("unit price"),
                        "total": row_data.get("total"),
                    }
                )
        return budget_data

    except (InvalidFileException, FileNotFoundError) as e:
        raise ExcelParsingError(f"Failed to open or read Excel file: {e}")
    except Exception as e:
        raise ExcelParsingError(f"An unexpected error occurred during Excel budget parsing: {e}")


# Example Usage (for testing purposes, if run directly)
if __name__ == "__main__":
    # Create dummy Excel files for testing
    def create_dummy_schedule_excel(path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        ws.append(["Task", "Start Date", "End Date", "Duration"])
        ws.append(["Project Kick-off", "2026-01-10", "2026-01-11", 2])
        ws.append(["Requirement Gathering", "2026-01-12", "2026-01-15", 4])
        ws.append(["Design Phase", "2026-01-16", "2026-01-20", 5])
        wb.save(path)

    def create_dummy_budget_excel(path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Item", "Quantity", "Unit Price", "Total"])
        ws.append(["Labor", 100, 50, 5000])
        ws.append(["Materials", 200, 10.5, 2100])
        ws.append(["Equipment Rental", 10, 200, 2000])
        wb.save(path)

    schedule_file = Path("dummy_schedule.xlsx")
    budget_file = Path("dummy_budget.xlsx")

    create_dummy_schedule_excel(schedule_file)
    create_dummy_budget_excel(budget_file)

    print("--- Simulating Excel Parsing Tests ---")

    # Test schedule parsing
    print(f"\nParsing schedule from: {schedule_file.name}")
    try:
        schedule = parse_schedule_from_excel(schedule_file)
        print(f"  SUCCESS: Parsed {len(schedule)} schedule items.")
        print(f"  First item: {schedule[0]}")
    except ExcelParsingError as e:
        print(f"  ERROR: {e}")

    # Test budget parsing
    print(f"\nParsing budget from: {budget_file.name}")
    try:
        budget = parse_budget_from_excel(budget_file)
        print(f"  SUCCESS: Parsed {len(budget)} budget items.")
        print(f"  First item: {budget[0]}")
    except ExcelParsingError as e:
        print(f"  ERROR: {e}")

    # Clean up dummy files
    schedule_file.unlink()
    budget_file.unlink()
