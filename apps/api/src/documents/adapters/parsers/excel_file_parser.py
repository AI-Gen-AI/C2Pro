"""
Excel Parser Adapter.

This adapter provides functionality to parse schedule and budget data from Excel files
using the openpyxl library, encapsulating external library details.
"""

import re
import unicodedata
from numbers import Number
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


class ExcelParsingError(Exception):
    """Custom exception for Excel parsing errors."""

    pass


class BudgetRows(list[dict[str, Any]]):
    """TS-COH-BUD-RECON-004: budget line items with optional declared total."""

    def __init__(
        self,
        rows: list[dict[str, Any]] | None = None,
        *,
        stated_total: float | None = None,
    ) -> None:
        super().__init__(rows or [])
        self.stated_total = stated_total


class ExcelFileParser:
    """
    Adapter class for parsing Excel files.
    Encapsulates the logic specific to the Excel format and `openpyxl` library.
    """

    async def parse_schedule(self, file_path: Path) -> list[dict[str, Any]]:
        """
        Parses schedule data from a standard Excel file format.

        Assumes the first sheet contains the schedule with a header row.
        Expected columns (case-insensitive): 'Task', 'Start Date', 'End Date', 'Duration'.

        Args:
            file_path: The path to the Excel file (.xlsx).

        Returns:
            A list of dictionaries, where each dictionary represents a schedule task.

        Raises:
            ExcelParsingError: If the file cannot be opened or does not match the expected format.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            header_row_index, headers = self._find_schedule_headers(sheet)
            if header_row_index is None:
                expected_headers = ["task/actividad", "start date/inicio", "end date/fin"]
                raise ExcelParsingError(
                    f"Missing one or more required headers: {', '.join(expected_headers)}"
                )

            schedule_data = []
            for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
                row_data = dict(zip(headers, row, strict=False))
                # Filter out empty rows
                if any(row_data.values()):
                    schedule_data.append(
                        {
                            "task": row_data.get("task"),
                            "start_date": row_data.get("start date"),
                            "end_date": row_data.get("end date"),
                            "duration": row_data.get("duration"),
                            "wbs": row_data.get("wbs"),
                            "predecessors": row_data.get("predecessors"),
                        }
                    )
            return schedule_data

        except ExcelParsingError:
            raise
        except (InvalidFileException, FileNotFoundError) as e:
            raise ExcelParsingError(f"Failed to open or read Excel file: {e}")
        except Exception as e:
            raise ExcelParsingError(
                f"An unexpected error occurred during Excel schedule parsing: {e}"
            )
        finally:
            workbook_to_close = locals().get("workbook")
            if workbook_to_close is not None and hasattr(workbook_to_close, "close"):
                workbook_to_close.close()

    @staticmethod
    def _schedule_header_aliases() -> dict[str, set[str]]:
        return {
            "task": {"task", "actividad", "tarea", "activity"},
            "start date": {"start date", "inicio", "comienzo", "fecha inicio"},
            "end date": {"end date", "fin", "termino", "fecha fin"},
            "duration": {"duration", "duracion"},
            "wbs": {"wbs"},
            "predecessors": {"predecessors", "predecesoras"},
        }

    @classmethod
    def _find_schedule_headers(cls, sheet: Any) -> tuple[int | None, list[str]]:
        aliases = cls._schedule_header_aliases()
        required = {"task", "start date", "end date"}

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            canonical_headers = cls._canonicalize_headers(row, aliases)
            if required.issubset(set(canonical_headers)):
                return row_index, canonical_headers

        first_row: Any = getattr(sheet, "__getitem__", lambda _: [])(1)
        fallback_headers = [
            getattr(cell, "value", cell) if cell is not None else "" for cell in first_row
        ]
        fallback_canonical_headers = cls._canonicalize_headers(fallback_headers, aliases)
        if required.issubset(set(fallback_canonical_headers)):
            return 1, fallback_canonical_headers

        return None, []

    async def parse_budget(self, file_path: Path) -> list[dict[str, Any]]:
        """
        Parses budget data from a standard Excel file format.

        Scans the first sheet for a budget header row.
        Expected canonical columns: 'Item', 'Quantity', 'Unit Price', 'Total'.

        Args:
            file_path: The path to the Excel file (.xlsx).

        Returns:
            A list of dictionaries, where each dictionary represents a budget line item.

        Raises:
            ExcelParsingError: If the file cannot be opened or does not match the expected format.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            header_row_index, headers = self._find_budget_headers(sheet)
            if header_row_index is None:
                expected_headers = ["item", "quantity", "unit price", "total"]
                accepted_aliases = "; ".join(
                    f"{canonical}: {', '.join(sorted(aliases))}"
                    for canonical, aliases in self._budget_header_aliases().items()
                )
                raise ExcelParsingError(
                    "Missing one or more required headers: "
                    f"{', '.join(expected_headers)}. Accepted aliases: {accepted_aliases}"
                )

            budget_data: BudgetRows = BudgetRows()
            for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
                row_data = dict(zip(headers, row, strict=False))
                if not any(row_data.values()):
                    continue
                stated_total = self._extract_declared_total(row_data)
                if stated_total is not None:
                    budget_data.stated_total = stated_total
                    continue
                quantity = self._coerce_budget_number(row_data.get("quantity"))
                unit_price = self._coerce_budget_number(row_data.get("unit price"))
                total = self._coerce_budget_number(row_data.get("total"))
                # Skip structural chapter/section subtotal rows. In a hierarchical
                # budget (presupuesto por capítulos) chapters carry a rolled-up total
                # with no quantity or unit price; emitting them as line items would
                # double-count against their own leaf partidas. A genuine line item
                # has a measured quantity and/or a unit price.
                if quantity is None and unit_price is None:
                    continue
                budget_data.append(
                    {
                        "item": row_data.get("item"),
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "total": total,
                    }
                )
            return budget_data

        except ExcelParsingError:
            raise
        except (InvalidFileException, FileNotFoundError) as e:
            raise ExcelParsingError(f"Failed to open or read Excel file: {e}")
        except Exception as e:
            raise ExcelParsingError(
                f"An unexpected error occurred during Excel budget parsing: {e}"
            )
        finally:
            workbook_to_close = locals().get("workbook")
            if workbook_to_close is not None and hasattr(workbook_to_close, "close"):
                workbook_to_close.close()

    @staticmethod
    def _budget_header_aliases() -> dict[str, set[str]]:
        return {
            "item": {
                "item",
                "partida",
                "descripción",
                "descripcion",
                "concepto",
                "código",
                "codigo",
                "capítulo",
                "capitulo",
            },
            "quantity": {
                "quantity",
                "cantidad",
                "medición",
                "medicion",
                "cant",
                "cant.",
                "uds",
                "ud",
                "unidades",
            },
            "unit price": {
                "unit price",
                "precio unitario",
                "precio",
                "p. unitario",
                "precio ud",
                "precio/ud",
                "importe unitario",
                "coste unitario",
            },
            "total": {
                "total",
                "importe",
                "importe total",
                "total partida",
                "subtotal",
                "coste",
            },
        }

    @classmethod
    def _find_budget_headers(cls, sheet: Any) -> tuple[int | None, list[str]]:
        aliases = cls._budget_header_aliases()
        required = {"item", "quantity", "unit price", "total"}

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            canonical_headers = cls._canonicalize_headers(row, aliases)
            if required.issubset(set(canonical_headers)):
                return row_index, canonical_headers

        first_row: Any = getattr(sheet, "__getitem__", lambda _: [])(1)
        fallback_headers = [
            getattr(cell, "value", cell) if cell is not None else "" for cell in first_row
        ]
        fallback_canonical_headers = cls._canonicalize_headers(fallback_headers, aliases)
        if required.issubset(set(fallback_canonical_headers)):
            return 1, fallback_canonical_headers

        return None, []

    @staticmethod
    def _normalize_header(value: Any) -> str:
        """Normalize a header cell for tolerant matching.

        Lowercases, strips accents, drops parenthetical qualifiers (e.g. ``(€)``,
        ``(Semana)``, ``(días)``), and reduces any punctuation/currency/symbols to
        single spaces. This lets real-world headers like ``PRECIO UNIT. (€)`` or
        ``Inicio (Semana)`` match their canonical aliases (``precio``, ``inicio``).
        """
        if value is None:
            return ""
        text = unicodedata.normalize("NFKD", str(value))
        text = "".join(ch for ch in text if not unicodedata.combining(ch)).lower()
        text = re.sub(r"\([^)]*\)", " ", text)  # drop parenthetical qualifiers
        text = re.sub(r"[^a-z0-9]+", " ", text)  # currency/punctuation -> space
        return " ".join(text.split())

    @classmethod
    def _match_canonical(cls, header_norm: str, aliases: dict[str, set[str]]) -> str | None:
        """Return the canonical column for a normalized header, or ``None``.

        A header matches an alias when every token of the (normalized) alias is
        present in the header's tokens (subset match), so ``importe (€)`` matches
        ``importe`` and ``precio unit. (€)`` matches ``precio``. When several
        aliases match, the most specific one wins (more tokens, then longer),
        which keeps ``importe unitario`` (unit price) from being read as ``importe``
        (total). Iteration is sorted so ties resolve deterministically.
        """
        if not header_norm:
            return None
        header_tokens = set(header_norm.split())
        best_score: tuple[int, int] = (0, 0)
        best_canonical: str | None = None
        for canonical, accepted in aliases.items():
            for alias in sorted(accepted):
                alias_tokens = cls._normalize_header(alias).split()
                if alias_tokens and all(tok in header_tokens for tok in alias_tokens):
                    score = (len(alias_tokens), len("".join(alias_tokens)))
                    if score > best_score:
                        best_score = score
                        best_canonical = canonical
        return best_canonical

    @classmethod
    def _canonicalize_headers(cls, row: Any, aliases: dict[str, set[str]]) -> list[str]:
        canonical: list[str] = []
        for value in row:
            header_norm = cls._normalize_header(value)
            match = cls._match_canonical(header_norm, aliases)
            canonical.append(match if match is not None else header_norm)
        return canonical

    @staticmethod
    def _coerce_budget_number(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, Number):
            return float(cast(float, value))
        if not isinstance(value, str):
            return None

        text = value.strip()
        if not text:
            return None

        cleaned = re.sub(r"[^0-9,.\-]", "", text)
        if cleaned in {"", "-", ".", ","}:
            return None

        if "," in cleaned and "." in cleaned:
            if cleaned.rfind(",") > cleaned.rfind("."):
                normalized = cleaned.replace(".", "").replace(",", ".")
            else:
                normalized = cleaned.replace(",", "")
        elif "," in cleaned:
            left, right = cleaned.rsplit(",", 1)
            if len(right) in {1, 2}:
                normalized = f"{left.replace(',', '')}.{right}"
            else:
                normalized = cleaned.replace(",", "")
        else:
            normalized = cleaned

        try:
            return float(normalized)
        except ValueError:
            return None

    @classmethod
    def _extract_declared_total(cls, row_data: dict[str, Any]) -> float | None:
        item = row_data.get("item")
        quantity = row_data.get("quantity")
        unit_price = row_data.get("unit price")
        total = cls._coerce_budget_number(row_data.get("total"))
        if total is None or quantity not in (None, "") or unit_price not in (None, ""):
            return None
        if not isinstance(item, str):
            return None
        label = item.strip().lower()
        declared_total_labels = (
            "total presupuesto",
            "total general",
            "total licitacion",
            "total licitación",
            "presupuesto total",
            "grand total",
            "declared total",
        )
        if label == "total" or any(marker in label for marker in declared_total_labels):
            return total
        return None
