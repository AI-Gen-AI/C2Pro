"""
Service for parsing complex Excel budget files.

This module provides a robust parser for .xlsx budget files, which often have
inconsistent formatting, such as dynamic header positions and varied column names.
"""

import io
import re
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Assuming the BOMItemBase is the canonical schema for a budget item.
# This aligns with the prompt's reference to schemas/bom.py.
from src.modules.projects.schemas import BOMItemBase as BudgetItem


# --- Constants for Parsing Logic ---

# Keywords to identify the header row. A row must contain at least one from each list.
HEADER_ROW_KEYWORDS = [
    {"descripción", "description", "concepto", "activity"},
    {"precio", "price", "importe", "total", "cost", "amount"},
]

# Mapping of potential column names to our canonical BudgetItem fields
COLUMN_ALIASES = {
    "item_code": ["item", "código", "codigo", "code", "partida"],
    "item_name": ["descripción", "descripcion", "description", "concepto", "activity"],
    "unit": ["unidad", "unit", "ud", "uom"],
    "quantity": ["cantidad", "quantity", "qty", "uds", "unidades"],
    "unit_price": ["precio unitario", "p.u.", "pu", "unit price", "precio u."],
    "total_price": ["importe", "total", "precio total", "total price", "cost", "amount"],
}


class ExcelBudgetParser:
    """
    Parses budget data from complex and loosely structured Excel files.
    """

    def parse_budget(self, file_content: bytes) -> Dict[str, Any]:
        """
        Main entry point for parsing an Excel file's binary content.

        Args:
            file_content: The binary content of the .xlsx file.

        Returns:
            A dictionary containing processed sheets count, a list of budget items,
            and the calculated total amount.
        """
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_content), read_only=True, data_only=True
            )
        except Exception as e:
            # Consider a more specific exception for file format errors
            raise ValueError(f"Failed to load Excel workbook: {e}")

        all_items: List[BudgetItem] = []
        sheets_processed = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Skip sheets that seem irrelevant
            if "portada" in sheet_name.lower() or "summary" in sheet_name.lower():
                continue

            sheet_items = self._parse_sheet(sheet)
            if sheet_items:
                all_items.extend(sheet_items)
                sheets_processed += 1
        
        total_amount = sum(item.total_price for item in all_items if item.total_price is not None)

        return {
            "sheets_processed": sheets_processed,
            "items": all_items,
            "total_amount": total_amount,
        }

    def _parse_sheet(self, sheet: Worksheet) -> List[BudgetItem]:
        """Parses an individual worksheet to extract budget items."""
        header_info = self._find_header_row(sheet)
        if not header_info:
            return []

        header_row_index, header_values = header_info
        column_map = self._map_columns(header_values)

        # Ensure mandatory columns are present
        if "item_name" not in column_map or ("total_price" not in column_map and ("quantity" not in column_map or "unit_price" not in column_map)):
            return []

        items: List[BudgetItem] = []
        for row in sheet.iter_rows(min_row=header_row_index + 1):
            row_values = [cell.value for cell in row]

            # Skip empty or total rows
            if all(v is None for v in row_values):
                continue
            
            description_val = row_values[column_map['item_name']]
            if isinstance(description_val, str) and "total" in description_val.lower():
                continue
            if not description_val:
                continue

            # Safely extract and clean data
            item_data = {}
            for field, col_idx in column_map.items():
                raw_val = row_values[col_idx]
                if field in ["quantity", "unit_price", "total_price"]:
                    item_data[field] = self._clean_numeric(raw_val)
                else:
                    item_data[field] = str(raw_val).strip() if raw_val is not None else None
            
            # Ensure description and quantity are present
            if not item_data.get("item_name") or not item_data.get("quantity"):
                 continue

            # Calculate total_price if not present
            if item_data.get("total_price") is None:
                if item_data.get("quantity") is not None and item_data.get("unit_price") is not None:
                    item_data["total_price"] = item_data["quantity"] * item_data["unit_price"]

            # Create a BudgetItem, pydantic will handle validation
            try:
                items.append(BudgetItem(**item_data))
            except Exception:
                # Could log this error; for now, we just skip the invalid row
                continue

        return items

    def _find_header_row(self, sheet: Worksheet) -> Optional[Tuple[int, List[str]]]:
        """
        Finds the header row in a sheet by searching for keywords.

        Iterates through the first 20 rows and returns the first row that
        contains at least one keyword from each required group.
        """
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20)):
            row_values = [str(cell.value).lower().strip() for cell in row if cell.value]
            
            # Check if all keyword groups are represented in the row
            if all(any(key in val for val in row_values for key in group) for group in HEADER_ROW_KEYWORDS):
                return i + 1, [str(cell.value) for cell in row]
        return None

    def _map_columns(self, header_values: List[str]) -> Dict[str, int]:
        """
        Maps the columns in the header to the canonical BudgetItem fields.
        """
        column_map: Dict[str, int] = {}
        normalized_headers = [h.lower().strip() for h in header_values]

        for canonical_field, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                try:
                    # Find the first occurrence of an alias
                    idx = normalized_headers.index(alias)
                    if canonical_field not in column_map:
                         column_map[canonical_field] = idx
                except ValueError:
                    continue
        return column_map
    
    def _clean_numeric(self, value: Any) -> Optional[float]:
        """
        Safely converts a value to a float, handling currency symbols and separators.
        """
        if isinstance(value, (int, float)):
            return float(value)
        if not isinstance(value, str):
            return None
        
        # Remove currency symbols, thousands separators, and strip whitespace
        cleaned_value = re.sub(r"[€$£]", "", value).strip()
        # Handle decimal commas
        if ',' in cleaned_value and '.' in cleaned_value:
             # Assume dot is thousands separator if it comes before comma
             if cleaned_value.find('.') < cleaned_value.find(','):
                  cleaned_value = cleaned_value.replace('.', '')
             cleaned_value = cleaned_value.replace(',', '.')
        elif ',' in cleaned_value:
             cleaned_value = cleaned_value.replace(',', '.')

        try:
            return float(cleaned_value)
        except (ValueError, TypeError):
            return None

