"""
Test Suite: Excel Parser Additional Tests
Component: Documents Module - Excel Parser Adapter
Priority: P0

Additional tests to improve Excel parser coverage.
"""

import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook


class TestExcelParserSchedule:
    """Tests for ExcelFileParser.parse_schedule method."""

    @pytest.mark.asyncio
    async def test_parse_schedule_success(self):
        """Test successful schedule parsing."""
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        parser = ExcelFileParser()

        # Create mock workbook
        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        mock_sheet.__getitem__ = MagicMock(
            side_effect=[
                [
                    MagicMock(value="Task"),
                    MagicMock(value="Start Date"),
                    MagicMock(value="End Date"),
                    MagicMock(value="Duration"),
                ],
                [
                    MagicMock(value="Task 1"),
                    MagicMock(value="2024-01-01"),
                    MagicMock(value="2024-01-31"),
                    MagicMock(value=30),
                ],
                [
                    MagicMock(value="Task 2"),
                    MagicMock(value="2024-02-01"),
                    MagicMock(value="2024-02-28"),
                    MagicMock(value=27),
                ],
            ]
        )
        mock_sheet.iter_rows = MagicMock(
            return_value=[
                ["Task 1", "2024-01-01", "2024-01-31", 30],
                ["Task 2", "2024-02-01", "2024-02-28", 27],
            ]
        )
        mock_workbook.active = mock_sheet

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            return_value=mock_workbook,
        ):
            # Create temp file path (not actual file)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                result = await parser.parse_schedule(Path(tmp_path))
                assert isinstance(result, list)
            finally:
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_parse_schedule_missing_headers(self):
        """Test schedule parsing with missing required headers."""
        from src.documents.adapters.parsers.excel_file_parser import (
            ExcelFileParser,
            ExcelParsingError,
        )

        parser = ExcelFileParser()

        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        mock_sheet.__getitem__ = MagicMock(
            side_effect=[
                [MagicMock(value="Task"), MagicMock(value="Wrong")],  # Missing Start Date, End Date
            ]
        )
        mock_workbook.active = mock_sheet

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            return_value=mock_workbook,
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                with pytest.raises(ExcelParsingError) as exc_info:
                    await parser.parse_schedule(Path(tmp_path))
                assert "Missing one or more required headers" in str(exc_info.value)
            finally:
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_parse_schedule_invalid_file(self):
        """Test schedule parsing with invalid file."""
        from src.documents.adapters.parsers.excel_file_parser import (
            ExcelFileParser,
            ExcelParsingError,
        )

        parser = ExcelFileParser()

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            side_effect=FileNotFoundError("File not found"),
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                with pytest.raises(ExcelParsingError):
                    await parser.parse_schedule(Path(tmp_path))
            finally:
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_parse_schedule_discovers_spanish_header_row_after_title_block(self):
        """TS-UD-DOC-XLS-001: realistic Spanish schedules parse after title rows."""
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        parser = ExcelFileParser()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "1. CRONOGRAMA DETALLADO"
        sheet.append([])
        sheet.append([])
        sheet.append(["ID", "WBS", "Actividad", "Duración (días)", "Inicio", "Fin", "Predecesoras"])
        sheet.append(["1.1", None, "Firma del contrato", 1, "2024-01-01", "2024-01-01", "–"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        workbook.save(tmp_path)
        workbook.close()

        result = await parser.parse_schedule(Path(tmp_path))

        assert result == [
            {
                "task": "Firma del contrato",
                "start_date": "2024-01-01",
                "end_date": "2024-01-01",
                "duration": 1,
                "wbs": None,
                "predecessors": "–",
            }
        ]


class TestExcelParserBudget:
    """Tests for ExcelFileParser.parse_budget method."""

    @pytest.mark.asyncio
    async def test_parse_budget_success(self):
        """Test successful budget parsing."""
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        parser = ExcelFileParser()

        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        # Header row
        header_cells = [
            MagicMock(value="Item"),
            MagicMock(value="Quantity"),
            MagicMock(value="Unit Price"),
            MagicMock(value="Total"),
        ]
        mock_sheet.__getitem__ = MagicMock(return_value=header_cells)
        mock_sheet.iter_rows = MagicMock(
            return_value=[
                ["Item 1", 10, 100, 1000],
                ["Item 2", 5, 200, 1000],
            ]
        )
        mock_workbook.active = mock_sheet

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            return_value=mock_workbook,
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                result = await parser.parse_budget(Path(tmp_path))
                assert isinstance(result, list)
            finally:
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_parse_budget_missing_headers(self):
        """Test budget parsing with missing required headers."""
        from src.documents.adapters.parsers.excel_file_parser import (
            ExcelFileParser,
            ExcelParsingError,
        )

        parser = ExcelFileParser()

        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        mock_sheet.__getitem__ = MagicMock(
            side_effect=[
                [MagicMock(value="Item")],  # Missing Quantity, Unit Price, Total
            ]
        )
        mock_workbook.active = mock_sheet

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            return_value=mock_workbook,
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                with pytest.raises(ExcelParsingError) as exc_info:
                    await parser.parse_budget(Path(tmp_path))
                assert "Missing one or more required headers" in str(exc_info.value)
            finally:
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_parse_budget_invalid_file(self):
        """Test budget parsing with invalid file."""
        from src.documents.adapters.parsers.excel_file_parser import (
            ExcelFileParser,
            ExcelParsingError,
        )

        parser = ExcelFileParser()

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            side_effect=FileNotFoundError("File not found"),
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                with pytest.raises(ExcelParsingError):
                    await parser.parse_budget(Path(tmp_path))
            finally:
                os.unlink(tmp_path)


class TestExcelParserEdgeCases:
    """Edge case tests for ExcelFileParser."""

    @pytest.mark.asyncio
    async def test_parse_schedule_empty_rows(self):
        """Test schedule parsing with empty rows."""
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        parser = ExcelFileParser()

        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        header_cells = [
            MagicMock(value="Task"),
            MagicMock(value="Start Date"),
            MagicMock(value="End Date"),
            MagicMock(value="Duration"),
        ]
        mock_sheet.__getitem__ = MagicMock(return_value=header_cells)
        mock_sheet.iter_rows = MagicMock(
            return_value=[
                ["Task 1", "2024-01-01", "2024-01-31", 30],
                [None, None, None, None],  # Empty row
                ["Task 2", "2024-02-01", "2024-02-28", 27],
            ]
        )
        mock_workbook.active = mock_sheet

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            return_value=mock_workbook,
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                result = await parser.parse_schedule(Path(tmp_path))
                # Empty rows should be filtered out
                assert len(result) == 2
            finally:
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_parse_budget_empty_rows(self):
        """Test budget parsing with empty rows."""
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        parser = ExcelFileParser()

        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        header_cells = [
            MagicMock(value="Item"),
            MagicMock(value="Quantity"),
            MagicMock(value="Unit Price"),
            MagicMock(value="Total"),
        ]
        mock_sheet.__getitem__ = MagicMock(return_value=header_cells)
        mock_sheet.iter_rows = MagicMock(
            return_value=[
                ["Item 1", 10, 100, 1000],
                [None, None, None, None],  # Empty row
                ["Item 2", 5, 200, 1000],
            ]
        )
        mock_workbook.active = mock_sheet

        with patch(
            "src.documents.adapters.parsers.excel_file_parser.openpyxl.load_workbook",
            return_value=mock_workbook,
        ):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            try:
                result = await parser.parse_budget(Path(tmp_path))
                # Empty rows should be filtered out
                assert len(result) == 2
            finally:
                os.unlink(tmp_path)


class TestExcelParserRealWorldHeaders:
    """TS-UD-DOC-XLS-002: tolerant header matching for real client Excel files.

    Regression for client budgets/schedules whose headers carry currency
    symbols, unit qualifiers, or combined labels (e.g. ``PRECIO UNIT. (€)``,
    ``Inicio (Semana)``, ``CAPÍTULO Y PARTIDA``). Exact-equality matching
    rejected these; normalized token matching accepts them.
    """

    async def _parse_budget(self, rows: list[list[object]]):
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        workbook.save(tmp_path)
        workbook.close()
        try:
            return await ExcelFileParser().parse_budget(Path(tmp_path))
        finally:
            os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_budget_accepts_decorated_spanish_headers(self):
        """Currency/abbreviation-decorated Spanish headers still map to canonicals."""
        result = await self._parse_budget(
            [
                [
                    "CÓDIGO",
                    "CAPÍTULO Y PARTIDA",
                    "UNIDAD",
                    "CANTIDAD",
                    "PRECIO UNIT. (€)",
                    "IMPORTE (€)",
                ],
                ["1.1.1", "Bombas sumergibles", "ud", 3, 45000, 135000],
                ["1.1.2", "Motores eléctricos", "ud", 3, 12000, 36000],
            ]
        )
        assert len(result) == 2
        assert result[0]["quantity"] == 3.0
        assert result[0]["unit_price"] == 45000.0
        assert result[0]["total"] == 135000.0

    @pytest.mark.asyncio
    async def test_budget_excludes_chapter_subtotal_rows(self):
        """Hierarchical chapter rows (total only, no qty/price) are not line items."""
        result = await self._parse_budget(
            [
                ["Código", "Partida", "Unidad", "Cantidad", "Precio Unit. (€)", "Importe (€)"],
                ["1", "MATERIALES", None, None, None, 171000],  # chapter subtotal -> skip
                ["1.1", "Bombas", "ud", 3, 45000, 135000],
                ["1.2", "Motores", "ud", 3, 12000, 36000],
            ]
        )
        # Only the two leaf partidas, not the chapter subtotal, are emitted.
        assert len(result) == 2
        assert sum(r["total"] for r in result) == 171000.0

    @pytest.mark.asyncio
    async def test_budget_importe_unitario_maps_to_unit_price_not_total(self):
        """Specificity: 'Importe unitario' is unit price, 'Importe' is total."""
        result = await self._parse_budget(
            [
                ["Partida", "Cantidad", "Importe unitario", "Importe"],
                ["Tubería", 10, 420, 4200],
            ]
        )
        assert len(result) == 1
        assert result[0]["unit_price"] == 420.0
        assert result[0]["total"] == 4200.0

    @pytest.mark.asyncio
    async def test_budget_extracts_unit_of_measure(self):
        """The UNIDAD (unit-of-measure) column is captured onto each line item."""
        result = await self._parse_budget(
            [
                ["CÓDIGO", "CAPÍTULO Y PARTIDA", "UNIDAD", "CANTIDAD", "PRECIO UNIT. (€)", "IMPORTE (€)"],
                ["1.1.1", "Bombas sumergibles", "ud", 3, 45000, 135000],
                ["2.2.1", "Soldadores homologados", "mes/h", "6 sold x 6m", "2.800,00/mes", 100800],
            ]
        )
        assert len(result) == 2
        assert result[0]["unit"] == "ud"
        assert result[1]["unit"] == "mes/h"
        # quantity/unit_price still parse (a real budget the parser already handled).
        assert result[0]["quantity"] == 3.0

    @pytest.mark.asyncio
    async def test_budget_infers_category_from_unit_and_name(self):
        """Category: time-priced labor -> service, equipment name -> equipment, measured -> material."""
        result = await self._parse_budget(
            [
                ["Partida", "Unidad", "Cantidad", "Precio Unit. (€)", "Importe (€)"],
                ["Bombas sumergibles centrifugas", "ud", 3, 45000, 135000],
                ["Tubería DN 300 (impulsión)", "m", 80, 420, 33600],
                ["Soldadores homologados", "mes/h", 6, 2800, 100800],
            ]
        )
        cats = {r["item"]: r.get("category") for r in result}
        assert cats["Bombas sumergibles centrifugas"] == "equipment"
        assert cats["Tubería DN 300 (impulsión)"] == "material"
        assert cats["Soldadores homologados"] == "service"

    @pytest.mark.asyncio
    async def test_budget_without_unit_column_still_parses(self):
        """A budget with no UNIDAD column still parses; unit is None (backward compatible)."""
        result = await self._parse_budget(
            [
                ["Partida", "Cantidad", "Precio Unit. (€)", "Importe (€)"],
                ["Tubería DN 300", 80, 420, 33600],
            ]
        )
        assert len(result) == 1
        assert result[0].get("unit") is None

    @pytest.mark.asyncio
    async def test_schedule_accepts_decorated_week_headers(self):
        """Schedule headers like 'Actividad / Tarea' and 'Inicio (Semana)' match."""
        from src.documents.adapters.parsers.excel_file_parser import ExcelFileParser

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            ["Nº", "Actividad / Tarea", "Duración (semanas)", "Inicio (Semana)", "Fin (Semana)"]
        )
        sheet.append([1, "Movilización", 2, 1, 2])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        workbook.save(tmp_path)
        workbook.close()
        try:
            result = await ExcelFileParser().parse_schedule(Path(tmp_path))
        finally:
            os.unlink(tmp_path)
        assert len(result) == 1
        assert result[0]["task"] == "Movilización"
        assert result[0]["start_date"] == 1
        assert result[0]["end_date"] == 2
        assert result[0]["duration"] == 2
